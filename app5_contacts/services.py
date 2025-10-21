import io
import csv
from openpyxl import load_workbook, Workbook
from django.http import HttpResponse

def parse_file(file):
    """Определяет формат файла и парсит в список словарей"""

    name = file.name.lower()
    if name.endswith('.csv'):
        return parse_csv(file)
    elif name.endswith('.xlsx'):
        return parse_xlsx(file)
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {file.name}")

def parse_csv(file):
    data_set = file.read().decode('utf-8')
    io_string = io.StringIO(data_set)
    reader = csv.DictReader(io_string, delimiter=',')
    return [row for row in reader]

def parse_xlsx(file):
    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return data

def prepare_contacts_for_import(contacts_file_data, existing_contacts, companies):
    """
    contacts_file_data - список словарей из CSV/XLSX
    existing_contacts - dict {'emails': {...}, 'phones': {...}}
    companies - dict {название -> id}
    """

    methods = []
    errors = []

    for row in contacts_file_data:
        first_name = str(row.get('имя', '')).strip()
        last_name = str(row.get('фамилия', '')).strip()
        phone_str = str(row.get('телефон', '')).strip() or str(row.get('номер телефона', '')).strip()
        email_str = str(row.get('email', '')).strip() or str(row.get('почта', '')).strip()
        company_name = str(row.get('компания', '')).strip()

        company_id = companies.get(company_name)
        if not company_id:
            errors.append(f"Компания '{company_name}' не найдена, контакт {first_name} {last_name} не создан.")
            continue

        # разбиваем телефоны и email на отдельные элементы
        phones = [{"VALUE": p.strip(), "VALUE_TYPE": "WORK"} for p in phone_str.split(",") if p.strip()]
        emails = [{"VALUE": e.strip(), "VALUE_TYPE": "WORK"} for e in email_str.split(",") if e.strip()]

        # проверка существующего контакта
        duplicate = False
        for e in emails:
            if e["VALUE"] in existing_contacts['emails']:
                duplicate = True
                break
        for p in phones:
            if p["VALUE"] in existing_contacts['phones']:
                duplicate = True
                break
        if duplicate:
            errors.append(f"Контакт {first_name} {last_name} с таким email или телефоном уже существует.")
            continue

        fields = {
            "NAME": first_name,
            "LAST_NAME": last_name,
            "PHONE": phones,
            "EMAIL": emails,
            "COMPANY_ID": company_id,
        }
        methods.append(("crm.contact.add", {"fields": fields}))

    return methods, errors


def contacts_to_csv_response(contacts, companies, filename='contacts.csv'):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(['имя', 'фамилия', 'телефон', 'email', 'компания'])
    for contact in contacts:
        name = contact.get('NAME', '')
        last_name = contact.get('LAST_NAME', '')
        phones = ', '.join([p.get('VALUE', '') for p in contact.get('PHONE', [])])
        emails = ', '.join([e.get('VALUE', '') for e in contact.get('EMAIL', [])])
        company_name = companies.get(contact.get('COMPANY_ID'), '')
        writer.writerow([name, last_name, phones, emails, company_name])
    return response

def contacts_to_xlsx_response(contacts, companies, filename='contacts.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    headers = ['имя', 'фамилия', 'телефон', 'email', 'компания']
    ws.append(headers)
    for contact in contacts:
        name = contact.get('NAME', '')
        last_name = contact.get('LAST_NAME', '')
        phones = ', '.join([p.get('VALUE', '') for p in contact.get('PHONE', [])])
        emails = ', '.join([e.get('VALUE', '') for e in contact.get('EMAIL', [])])
        company_name = companies.get(contact.get('COMPANY_ID'), '')
        ws.append([name, last_name, phones, emails, company_name])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
